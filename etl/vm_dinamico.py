"""
vm_dinamico.py — VM Dinâmico + Pulmão por SKU

VM (prateleira): demanda média × cobertura × taxa × correção
Pulmão (armário): Z × σ_diária × √lead_time
"""

import pandas as pd
import numpy as np
import math
from pathlib import Path

NIVEL_SERVICO_Z = {90: 1.28, 95: 1.65, 97: 1.88, 98: 2.05, 99: 2.33}


def _nivel_para_z(nivel: float) -> float:
    if nivel <= 1:
        nivel = nivel * 100
    return NIVEL_SERVICO_Z.get(round(nivel), 1.65)


def carregar_parametros_vm(caminho_xlsx: str) -> dict:
    caminho = Path(caminho_xlsx)
    avisos = []

    if not caminho.exists():
        return {"ok": False, "erros": [f"Arquivo não encontrado: {caminho}"], "avisos": []}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(caminho), data_only=True)
    except Exception as e:
        return {"ok": False, "erros": [f"Erro ao ler: {e}"], "avisos": []}

    # Globais
    globais = {
        "dias_cobertura": 15, "inicio_alta": 10, "fim_alta": 3,
        "mult_pa": 2.0, "vm_minimo": 2, "lead_time": 3, "nivel_servico_default": 95,
    }
    if "Parametros_Globais" in wb.sheetnames:
        ws = wb["Parametros_Globais"]
        mapa = {
            "Dias Cobertura VM": "dias_cobertura", "Início Alta Temporada": "inicio_alta",
            "Fim Alta Temporada": "fim_alta", "Multiplicador PA": "mult_pa",
            "VM Mínimo Absoluto": "vm_minimo", "Lead Time Reposição": "lead_time",
            "Nível de Serviço Padrão": "nivel_servico_default",
        }
        for r in range(4, 20):
            p, v = ws.cell(r, 2).value, ws.cell(r, 3).value
            if p and str(p).strip() in mapa and v is not None:
                try:
                    globais[mapa[str(p).strip()]] = float(v)
                except (ValueError, TypeError):
                    avisos.append(f"Valor inválido para '{p}': {v}")

    # Colégios (A=nome, B=taxa, C=nível serviço)
    colegios = {}
    if "Parametros_Colegio" in wb.sheetnames:
        ws = wb["Parametros_Colegio"]
        for r in range(2, 100):
            nome_c, taxa, ns = ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value
            if nome_c and str(nome_c).strip():
                n = str(nome_c).strip()
                try:
                    t = float(taxa) if taxa is not None else 1.0
                except (ValueError, TypeError):
                    t = 1.0
                try:
                    s = float(ns) if ns is not None else globais["nivel_servico_default"]
                except (ValueError, TypeError):
                    s = globais["nivel_servico_default"]
                colegios[n] = {"taxa_crescimento": t, "nivel_servico": s}

    # SKUs (A=código, B=produto, C=correção)
    skus = {}
    if "Parametros_SKU" in wb.sheetnames:
        ws = wb["Parametros_SKU"]
        for r in range(2, 2000):
            cod, corr = ws.cell(r, 1).value, ws.cell(r, 3).value
            if cod and str(cod).strip():
                c = str(cod).strip()
                try:
                    skus[c] = float(corr) if corr is not None else 1.0
                except (ValueError, TypeError):
                    skus[c] = 1.0

    return {"globais": globais, "colegios": colegios, "skus": skus,
            "ok": True, "erros": [], "avisos": avisos}


def _filtrar_alta(df, inicio_mes, fim_mes):
    mes = df["Data"].dt.month
    if inicio_mes <= fim_mes:
        return df[(mes >= inicio_mes) & (mes <= fim_mes)]
    return df[(mes >= inicio_mes) | (mes <= fim_mes)]


def calcular_vm_por_sku(dados: dict, params: dict) -> dict:
    glob = params["globais"]
    map_colegios = params["colegios"]
    map_skus_correcao = params["skus"]

    dias_cobertura = glob["dias_cobertura"]
    inicio_alta, fim_alta = int(glob["inicio_alta"]), int(glob["fim_alta"])
    mult_pa = glob["mult_pa"]
    vm_minimo = int(glob["vm_minimo"])
    lead_time = glob["lead_time"]
    ns_default = glob["nivel_servico_default"]

    itens = dados["itens"]
    produtos = dados["produtos"]
    detalhes = dados["detalhes"]

    map_id_colegio = detalhes.set_index("ID_produto")["Marca_sku"].to_dict()

    if inicio_alta <= fim_alta:
        meses_alta = list(range(inicio_alta, fim_alta + 1))
    else:
        meses_alta = list(range(inicio_alta, 13)) + list(range(1, fim_alta + 1))
    dias_alta = len(meses_alta) * 30

    itens_alta = _filtrar_alta(itens, inicio_alta, fim_alta)

    # === Pré-cálculos vetorizados ===
    if len(itens_alta) > 0:
        agg_alta = itens_alta.groupby("ID_produto").agg(
            pecas=("Quantidade", "sum"), pedidos=("ID_pedido", "nunique")
        )

        # Vendas diárias por produto × dia
        itens_alta_c = itens_alta.copy()
        itens_alta_c["dia"] = itens_alta_c["Data"].dt.date
        vd_diario = itens_alta_c.groupby(["ID_produto", "dia"])["Quantidade"].sum()

        # Dict: id_prod → array de quantidades por dia COM venda
        _vd_grouped = vd_diario.groupby("ID_produto").apply(lambda x: x.values).to_dict()

        # Range de dias POR PRODUTO (corrige bug do σ global)
        _date_range = itens_alta_c.groupby("ID_produto")["dia"].agg(["min", "max"])
        _n_dias_por_prod = {}
        for id_p, row in _date_range.iterrows():
            _n_dias_por_prod[id_p] = (pd.Timestamp(row["max"]) - pd.Timestamp(row["min"])).days + 1
    else:
        agg_alta = pd.DataFrame(columns=["pecas", "pedidos"])
        _vd_grouped = {}
        _n_dias_por_prod = {}

    # === Loop (só dict lookups) ===
    resultado = {}

    for _, prod in produtos.iterrows():
        id_prod = str(prod["ID"]).strip()
        sku = prod["codigo"]

        pecas_alta = agg_alta.loc[id_prod, "pecas"] if id_prod in agg_alta.index else 0
        pedidos_alta = agg_alta.loc[id_prod, "pedidos"] if id_prod in agg_alta.index else 0

        d_alta = pecas_alta / dias_alta if dias_alta > 0 else 0
        pa = pecas_alta / pedidos_alta if pedidos_alta > 0 else 1.0
        pedidos_dia = pedidos_alta / dias_alta if dias_alta > 0 else 0

        # σ por SKU (range específico do produto, não global)
        qtds_dias = _vd_grouped.get(id_prod, np.array([]))
        n_dias_reais = _n_dias_por_prod.get(id_prod, dias_alta)
        if len(qtds_dias) > 0 and n_dias_reais > 1:
            dias_sem = max(0, n_dias_reais - len(qtds_dias))
            todas = np.concatenate([qtds_dias, np.zeros(dias_sem)])
            sigma = float(np.std(todas, ddof=1)) if len(todas) > 1 else 0.0
            if np.isnan(sigma):
                sigma = 0.0
        else:
            sigma = 0.0

        colegio = str(map_id_colegio.get(id_prod, "")).strip()
        col_p = map_colegios.get(colegio, {})
        taxa_cresc = col_p.get("taxa_crescimento", 1.0) if col_p else 1.0
        nivel_servico = col_p.get("nivel_servico", ns_default) if col_p else ns_default
        z = _nivel_para_z(nivel_servico)
        correcao = map_skus_correcao.get(sku, 1.0)

        # VM
        vm_cobertura = d_alta * dias_cobertura * taxa_cresc * correcao
        vm_piso = pa * mult_pa
        vm_bruto = max(vm_cobertura, vm_piso, vm_minimo)
        vm_final = math.ceil(vm_bruto)

        if vm_bruto <= vm_minimo:
            fonte = "minimo_absoluto"
        elif vm_piso >= vm_cobertura:
            fonte = "piso_PA"
        else:
            fonte = "cobertura"

        # Pulmão
        pulmao = math.ceil(z * sigma * math.sqrt(lead_time))
        total = vm_final + pulmao

        resultado[sku] = {
            "vm": vm_final, "pulmao": pulmao, "total": total,
            "d_alta": round(d_alta, 4), "sigma": round(sigma, 4),
            "pa": round(pa, 2), "pedidos_dia": round(pedidos_dia, 4),
            "taxa_cresc": taxa_cresc, "correcao": correcao,
            "colegio": colegio, "fonte_vm": fonte,
        }

    return resultado
